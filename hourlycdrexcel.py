###########################################
## Program: Hourly Call Reports          ##
## Description: Reports on # of inbound  ##
## calls from each hour of the day. Then ##
## takes averages the hours for each day ##
## separately.							 ##
## Date: 2/2/15							 ##
## Author: Jeffrey Zic					 ##
###########################################

import re
import fileinput
import sys
from datetime import date, timedelta
import os.path
import datetime
from collections import namedtuple
from openpyxl import Workbook
import openpyxl

fname = sys.stdin.readlines() # Receives from stdin, built to work with cdr.sh

inbound_group = ['"CCR South Front Desk"', '"Cindy Business Office"', '"Tom Back Office"', '"CCR North Front Desk"', '"CCR Workstation"', '"Nan Office Manager"', '"Library East"', '"Doctor Workstation South"', '"Pharmacy Counter"', '"Team Leader Station"', '"Library South West"', '"Back Office North"', '"Sandi CCR TL"', '"Chris P Back Office South"', '"Tx Room West"']

weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

def getYesterday():
	return date.today() - timedelta(1)
		
def getDay():
	return weekdays[getYesterday().weekday()]

class CDR: #There are other parameters available for the CDRs but I just added the ones that are needed

	CallDetailStruct = namedtuple("CallDetailStruct", "end_timestamp, direction destination_name, hangup_cause, caller_id_name, destination_type")
	
	CDR_List = []
	hourlycounts = [0]*24
	
	def createCharts():
	
		wb = load_workbook('CDR.xlsx')
		
		ws = wb.activesheet()
		sheetpos = 1
		
		for i in weekdays:
			ws.title = weekdays
			
			for j in range(2, 26):
				ws.cell(row = j, column = 1) = j-2
				
			ws = wb.create_sheet(sheetpos)
			sheetpos += 1
		
		ws = wb.create_sheet("Averages")
		
		for i in range(2,26):
			ws.cell(row = i, column = 1) = i - 2
			
		columnpos = 2
		for i in weekdays:
			ws.cell(row = 1, column = columnpos) = i
			columnpos += 1
		
		ws.cell(row = 1, column = 1) = columnpos
					
	def getCalls():
	# grabs calls from a file generated from the Cudatel's REST API, run in bash and filename grabbed from STDIN
		newCall = True
		
		with open(fname[0].rstrip('\n'),'r') as f:
		
			for line in f:
				
				if newCall == True: # In the records, the parameters always appear in the same order so a new call is determined by
					newCDR = CallDetailStruct # seeing when it reaches the last parameter in a call, then the next one will be of a new call
					newCall = False
			
				words = line.rstrip('\n').partition(":")
				type = words[0].lstrip(' ').rstrip(' ')
				data = words[-1].rstrip(',').rstrip(' ').lstrip(' ')
						
				if (type == '"end_timestamp"'):
					newCDR.end_timestamp = data
				elif (type == '"direction"'):
					newCDR.direction = data
				elif (type == '"destination_name"'):
					newCDR.destination_name = data
				elif (type == '"hangup_cause"'):
					newCDR.hangup_cause = data
				elif (type == '"caller_id_name"'):
					newCDR.caller_id_name = data
				elif (type == '"destination_type"'):
					newCDR.destination_type = data
					this.CDR_List.append(newCDR)
					newCall = True
					
	def	RecordTotals():
	#Record the total calls in a .xlsx file
	
		wb = load_workbook('CDR.xlsx')
		
		ws = wb.get_sheet_by_name('Totals')
		
		i = 0
		j = 1
				
		while ws.cell(row = i, column = j) not "":
			j += 1

		ws.cell(row = i, column = j) = date.today() - timedelta(1)
		
		z = 0
	
		for y in range(2, 26):
			ws.cell(row = y, column = j) = hourlycounts[z]
			z += 1
			
		wb.save('CDR.xlsx')
		
	def RecordAvgs():
		
		wb = load_workbook('CDR.xlsx')
		
		currentDate = getYesterday()
		currentDay = getDay()

		ws = wb.get_sheet_by_name("Averages")
		wsTotals = wb.get_sheet_by_name('Totals')
		days = wsTotals.cell(row = 1, column = 1)
		
		for day in weekdays:
			for j in range(2, 25):
		
				totalCalls = []
		
				for i in range(2, days + 1):
					if wsTotals.cell(row = 1, column = j).weekday() == currentDay: 
						totalCalls.append(wsTotals.cell(row = j, column = j))
		
				dayPos = weekdays.index(currentDay)
		
				avg = sum(totalCalls) / totalCalls.len()
		
				ws.cell(row = j, column = dayPos + 1) = avg
			
			wb.save('CDR.xlsx')
	
	def getTotalsRange1Hour(firstDay, lastDay, hour):
	
		wb = load_workbook('CDR.xlsx')
		ws = wb.get_sheet_by_name('Totals')
		
		totals = []
		
		foundFirst = ""
		foundLast = ""
		columnpos = 1
		
		while foundFirst not firstDay:
			foundFirst = ws.cell(row = 1, column = columnpos)
			columnpos += 1
		while foundLast not lastDay:
			totals.append(ws.cell(row = hour + 1, column = columnpos))
			columnpos += 1
		
		return totals
	
	def getTotalsRangeAllHour(firstDay, lastDay):
		wb = load_workbook('CDR.xlsx')
		ws = wb.get_sheet_by_name('Totals')
		
		totals = []
		
		foundFirst = ""
		foundLast = ""
		columnpos = 1
		
		while foundFirst not firstDay:
			foundFirst = ws.cell(row = 1, column = columnpos)
			columnpos += 1
		while foundLast not lastDay:
			currDay = []
			for hour in range(1,24):
				currDay.append(ws.cell(row = hour, column = columnpos))
			totals.append(currDay)
			columnpos += 1
		
		return totals
		
	def getAvgsRange(firstDay, lastDay):
		totals = getTotalsRangeAllHour(firstDay,lastDay)
		
	
	def createCharts():
			
	def CountCalls():
	# Counts the incoming calls received each hour.
		
		for call in this.CDR_List:
		
			if (call.direction == '"inbound"') and ((call.hangup_cause == '"NORMAL_CLEARING"') or (call.hangup_cause == '"NONE"')) and (not call.caller_id_name in inbound_group): #Filtering out any calls in the inbound group so we don't count any intra-office calls
				hourlycounts[int(call.end_timestamp.split()[1][:2])] += 1					
					
def transferF(old, new):
# for replacing a file of the same name
	os.remove(new)
	os.rename(old, new)
	
def dailyCalls():
# records the calls/ hr of each hour every day
	weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
	yesterdayDate = date.today() - timedelta(1)
		
	with open("./Daily/" + weekdays[yesterdayDate.weekday()], 'a') as f:
	
		f.write(str(yesterdayDate)[:10] + ": \n")
		for i in range(24):
			f.write(str(i) + ": " + str(hourlycounts[i]) + "\n")
		f.write("\n")
	
def AvgCalls():
# records the average amount of calls received each hour for each day of the week
	weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
	firstTime = False
	
	if not os.path.isfile("./Reports/Averages"): #if the file to store the data doesn't exist we need to create it with correct formatting
		
		for day in weekdays:
		
			with open("./Reports/Averages", 'a') as f:
				f.write(day + '\n')
				count = 0
				while count < 24:
					f.write(str(count).zfill(2) + " : -\n") #If a lines are marked with a dash that means the program has never been run for that day
					count += 1
				f.write("\n")
		
	with open("./Reports/Averages", 'r') as f, open("./Reports/tempavg", 'w') as o:
	
		currentDate = date.today() - timedelta(1)
		currentDay = weekdays[currentDate.weekday()]
		pastAvg = [None] * 24
		foundDay = 0
		count = 1
		
		for line in f:
			
			if line.rstrip('\n') == currentDay:
				foundDay = 1
				o.write(line)
				continue
					
			if foundDay == 1 and count <= 24:
				nums = line.split()
				currHour = nums[0]
				pastAvg[int(currHour)] = nums[2]
				if pastAvg[int(currHour)] == "-":
					newAvg = int(hourlycounts[int(currHour)])
				else:
					newAvg = (int(pastAvg[int(currHour)]) + int(hourlycounts[int(currHour)])) / 2
				o.write(str(currHour).zfill(2) + " : " + str(int(newAvg)) + "\n")
				count += 1
			else:
				o.write(line)

	transferF("./Reports/tempavg", "./Reports/Averages")
	
def main():	
	CountCalls()
	AvgCalls()
	dailyCalls()

if __name__ == "__main__":
	main()